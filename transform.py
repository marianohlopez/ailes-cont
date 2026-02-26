from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
import yagmail
from dotenv import load_dotenv
import os

# --- CARGAR VARIABLES DE ENTORNO ---
load_dotenv()

DB_HOST = os.getenv('DB_HOST')
DB_PORT = int(os.getenv('DB_PORT'))
DB_USER = os.getenv('DB_USER')
DB_PASS = os.getenv('DB_PASS')
DB_NAME = os.getenv('DB_NAME')
MAIL_AUTOR = os.getenv("MAIL_AUTOR")
APP_GMAIL_PASS = os.getenv("APP_GMAIL_PASS")
MAIL_DESTINO = os.getenv("MAIL_DESTINO")


def transformar_datos(registros, hoy, condicion):
    resultados = []
    for id, importe, fecha_str, fec_envio_os, descrip, periodo, oSocial, alum_nombre, alum_apellido, obs, etiqueta in registros:
        if fecha_str:
            try:
                fecha = fecha_str if isinstance(fecha_str, datetime) else datetime.strptime(str(fecha_str), '%Y-%m-%d')
                dias = (hoy - fecha).days
                if condicion == 'todas':
                    resultados.append([
                        id,
                        importe,
                        fecha.date(),
                        fec_envio_os,
                        dias,
                        descrip,
                        periodo,
                        oSocial,
                        f"{alum_apellido}, {alum_nombre}",
                        obs,
                        etiqueta
                    ])
                elif dias > 45:
                    resultados.append([
                        id,
                        importe,
                        fecha.date(),
                        fec_envio_os,
                        dias,
                        descrip,
                        periodo,
                        oSocial,
                        f"{alum_apellido}, {alum_nombre}",
                        "",
                        obs,
                        etiqueta
                    ])
            except Exception as e:
                print(f"Error procesando la fecha {fecha_str} para ID {id}: {e}")
    return resultados


def exportar_excel(datos_alertas, datos_alertas_todos, datos_cobrados, prest_sin_pa, hoy):
    wb = Workbook()
    ws = wb.active
    ws.title = "Alertas"

    headers_alertas = ["ID_Factura", "Importe", "Fecha de fact.", "Fecha envío OS", "Días desde fecha de fact.", 
                       "Estado", "Periodo", "OS", "Alumno", "A Indyco", "Observaciones", "Etiqueta"]
    ws.append(headers_alertas)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for fila in datos_alertas:
        ws.append(fila)

    # Segunda hoja: Facturas Cobradas Recientes
    ws2 = wb.create_sheet(title="Alertas (todas)")

    headers_alertas_todas = ["ID_Factura", "Importe", "Fecha de fact.", "Fecha envío OS", "Días desde fecha de fact.", 
                       "Estado", "Periodo", "OS", "Alumno", "Observaciones", "Etiqueta"]
    ["ID_Factura", "Importe", "Fecha de fact.", "Fecha envío OS", "Días desde fecha de fact.", 
                       "Estado", "Periodo", "OS", "Alumno", "A Indyco", "Observaciones", "Etiqueta"]
    
    ws2.append(headers_alertas_todas)

    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for fila in datos_alertas_todos:
        ws2.append(fila)

    # Tercera hoja: Facturas Cobradas Recientes
    ws3 = wb.create_sheet(title="Cobradas dentro de los 60 días")
    headers_cobradas = ["ID_Factura", "Importe", "Fecha de fact.", "Fecha envío OS", "Fecha de cobro", "Estado", 
                        "Periodo", "OS", "Alumno", "A Indyco", "Observaciones", "Etiqueta"]
    ws3.append(headers_cobradas)
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for id, importe, cbteFch, fec_envio, fec_cobro, estado, periodo, os, nom, ape, obs, etiqueta in datos_cobrados:
        ws3.append([
            id,
            importe,
            cbteFch.date() if isinstance(cbteFch, datetime) else cbteFch,
            fec_envio.date() if isinstance(fec_envio, datetime) else fec_envio,
            fec_cobro.date() if isinstance(fec_envio, datetime) else fec_cobro,
            estado,
            periodo,
            os,
            f"{ape}, {nom}",
            "",
            obs,
            etiqueta
        ])

    # Cuarta hoja: Prestaciones sin PA > 60 dias
    ws4 = wb.create_sheet(title="Prestaciones sin PA > 60 días")
    headers_sin_pa = ["PRESTACION ID", "ALUMNO", "FEC. DE ÚLTIMA BAJA", "DÍAS SIN PA"]
    ws4.append(headers_sin_pa)
    for cell in ws4[1]:
        cell.font = Font(bold=True)

    for row in prest_sin_pa:
        ws4.append(row)

    nombre_archivo = f"reporte_contable_{hoy.strftime('%Y-%m-%d')}.xlsx"
    wb.save(nombre_archivo)
    print(f"Archivo Excel generado: {nombre_archivo}")
    return nombre_archivo


def enviar_correo(nombre_archivo):
    try:
        yag = yagmail.SMTP(MAIL_AUTOR, APP_GMAIL_PASS)
        yag.send(
            to=MAIL_DESTINO,
            subject="Reporte de Facturas emitidas - Cobros",
            contents= """Buenos días, se adjunta el reporte semanal del área contable.
              \nSaludos,\nMariano López - Ailes Inclusión.""",
            attachments=nombre_archivo
        )
        print("Correo enviado correctamente.")
    except Exception as e:
        print("Error al enviar el correo:", e)