import mysql.connector
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
import yagmail
from dotenv import load_dotenv
import os

# --- CARGAR VARIABLES DE ENTORNO ---
load_dotenv()

DB_HOST = os.getenv('DB_HOST')
DB_PORT = int(os.getenv('DB_PORT'))
DB_USER = os.getenv('DB_USER')
DB_PASS = os.getenv('DB_PASS')
DB_NAME = os.getenv('DB_NAME')
MAIL_AUTOR = os.getenv("MAIL_AUTOR")
APP_GMAIL_PASS = os.getenv("APP_GMAIL_PASS")
MAIL_DESTINO = os.getenv("MAIL_DESTINO")


def conectar_db():
    try:
        conn = mysql.connector.connect(
            host=DB_HOST,
            port=DB_PORT,
            user=DB_USER,
            password=DB_PASS,
            database=DB_NAME
        )
        print("Conexión a la base de datos exitosa.")
        return conn
    except Exception as e:
        print("Error al conectar a la base de datos:", e)
        exit(1)


def extraer_datos_deudas(cursor):
    query = """
        SELECT c.NroComprobante, c.ImpTotal, c.cbteFch, c.fec_envio_os, c.factura_cobro_descrip, c.mes_anio, 
            o.os_nombre, p.alumno_nombre, p.alumno_apellido, c.factura_obs, e.etiqueta
        FROM v_comprobantes c
        LEFT JOIN v_etiquetas_facturas e ON c.id = e.comprobante_id
        JOIN v_os o ON c.os_id = o.os_id
        JOIN v_prestaciones p ON c.prestacion_id = p.prestacion_id
        WHERE YEAR(cbteFch) = 2025
          AND factura_cobro_descrip = 'PENDIENTE' COLLATE utf8mb4_0900_ai_ci
    """
    cursor.execute(query)
    return cursor.fetchall()

def extraer_datos_cobrados(cursor):
    query = """
        SELECT c.NroComprobante, c.ImpTotal, c.cbteFch, c.fec_envio_os,  c.cobro_fec, c.factura_cobro_descrip, 
            c.mes_anio, o.os_nombre, p.alumno_nombre, p.alumno_apellido, c.factura_obs, e.etiqueta
        FROM v_comprobantes c
        LEFT JOIN v_etiquetas_facturas e ON c.id = e.comprobante_id
        JOIN v_os o ON c.os_id = o.os_id
        JOIN v_prestaciones p ON c.prestacion_id = p.prestacion_id
        WHERE c.fec_envio_os IS NOT NULL
            AND YEAR(cbteFch) = 2025
            AND (
                c.factura_cobro_descrip = 'COBRADA TOTAL' COLLATE utf8mb4_0900_ai_ci OR
                c.factura_cobro_descrip = 'COBRADA PARCIAL' COLLATE utf8mb4_0900_ai_ci
            )
            AND STR_TO_DATE(c.cobro_fec, '%Y-%m-%d') 
                BETWEEN STR_TO_DATE(c.fec_envio_os, '%Y-%m-%d')
                    AND DATE_ADD(STR_TO_DATE(c.fec_envio_os, '%Y-%m-%d'), INTERVAL 60 DAY);
    """
    cursor.execute(query)
    return cursor.fetchall()

def extract_prest_sin_pa(cursor):
  query = """ 
    SELECT 
      p.prestacion_id,
      CONCAT(p.alumno_apellido, ", ",p.alumno_nombre) as alumno_completo,
      DATE_FORMAT(COALESCE(MAX(a.asignpa_pa_fec_baja), a.asignpa_fec1), '%d-%m%-%Y') AS ultima_fecha_sin_pa,
      DATEDIFF(CURDATE(), COALESCE(MAX(a.asignpa_pa_fec_baja), a.asignpa_fec1)) AS dias_sin_pa
    FROM 
      v_prestaciones p
    LEFT JOIN 
      v_asignaciones_pa a 
      ON p.prestacion_id = a.asignpa_prest
    WHERE 
      p.prestipo_nombre_corto != 'TERAPIAS'
      AND p.prestacion_pa IS NULL
      AND p.prestacion_estado = 1
      AND p.prestacion_alumno != 522
    GROUP BY 
      p.prestacion_id, p.prestacion_alumno
    HAVING 
      dias_sin_pa > 60
	  ORDER BY dias_sin_pa;
    """
  cursor.execute(query)
  return cursor.fetchall()

def transformar_datos(registros, hoy):
    resultados = []
    for id, importe, fecha_str, fec_envio_os, descrip, periodo, oSocial, alum_nombre, alum_apellido, obs, etiqueta in registros:
        if fecha_str:
            try:
                fecha = fecha_str if isinstance(fecha_str, datetime) else datetime.strptime(str(fecha_str), '%Y-%m-%d')
                dias = (hoy - fecha).days
                if dias > 45:
                    resultados.append([
                        id,
                        importe,
                        fecha.date(),
                        fec_envio_os,
                        dias,
                        descrip,
                        periodo,
                        oSocial,
                        f"{alum_apellido}, {alum_nombre}",
                        "",
                        obs,
                        etiqueta
                    ])
            except Exception as e:
                print(f"Error procesando la fecha {fecha_str} para ID {id}: {e}")
    return resultados


def exportar_excel(datos_alertas, datos_cobrados, prest_sin_pa, hoy):
    wb = Workbook()
    ws = wb.active
    ws.title = "Alertas"

    headers_alertas = ["ID_Factura", "Importe", "Fecha de fact.", "Fecha envío OS", "Días desde fecha de fact.", 
                       "Estado", "Periodo", "OS", "Alumno", "A Indyco", "Observaciones", "Etiqueta"]
    ws.append(headers_alertas)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for fila in datos_alertas:
        ws.append(fila)

    # Segunda hoja: Facturas Cobradas Recientes
    ws2 = wb.create_sheet(title="Cobradas dentro de los 60 días")
    headers_cobradas = ["ID_Factura", "Importe", "Fecha de fact.", "Fecha envío OS", "Fecha de cobro", "Estado", 
                        "Periodo", "OS", "Alumno", "A Indyco", "Observaciones", "Etiqueta"]
    ws2.append(headers_cobradas)
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for id, importe, cbteFch, fec_envio, fec_cobro, estado, periodo, os, nom, ape, obs, etiqueta in datos_cobrados:
        ws2.append([
            id,
            importe,
            cbteFch.date() if isinstance(cbteFch, datetime) else cbteFch,
            fec_envio.date() if isinstance(fec_envio, datetime) else fec_envio,
            fec_cobro.date() if isinstance(fec_envio, datetime) else fec_cobro,
            estado,
            periodo,
            os,
            f"{ape}, {nom}",
            "",
            obs,
            etiqueta
        ])

    # Tercera hoja: Prestaciones sin PA > 60 dias
    ws3 = wb.create_sheet(title="Prestaciones sin PA > 60 días")
    headers_sin_pa = ["PRESTACION ID", "ALUMNO", "FEC. DE ÚLTIMA BAJA", "DÍAS SIN PA"]
    ws3.append(headers_sin_pa)
    for cell in ws3[1]:
        cell.font = Font(bold=True)

    for row in prest_sin_pa:
        ws3.append(row)

    nombre_archivo = f"reporte_contable_{hoy.strftime('%Y-%m-%d')}.xlsx"
    wb.save(nombre_archivo)
    print(f"Archivo Excel generado: {nombre_archivo}")
    return nombre_archivo


def enviar_correo(nombre_archivo):
    try:
        yag = yagmail.SMTP(MAIL_AUTOR, APP_GMAIL_PASS)
        yag.send(
            to=MAIL_DESTINO,
            subject="Reporte de Facturas emitidas - Cobros",
            contents= """Buenos días, se adjunta el reporte semanal del área contable.
              \nSaludos,\nMariano López - Ailes Inclusión.""",
            attachments=nombre_archivo
        )
        print("Correo enviado correctamente.")
    except Exception as e:
        print("Error al enviar el correo:", e)


def main():
    hoy = datetime.now()

    conn = conectar_db()
    cursor = conn.cursor()

    registros_alertas = extraer_datos_deudas(cursor)
    print(f"Registros de alerta extraídos: {len(registros_alertas)}")

    datos_alertas = transformar_datos(registros_alertas, hoy)

    registros_cobrados = extraer_datos_cobrados(cursor)
    print(f"Registros de cobradas recientes: {len(registros_cobrados)}")

    prest_sin_pa = extract_prest_sin_pa(cursor)
    print(f"Registros de prestaciones sin pa > 60 días: {len(prest_sin_pa)}")

    if datos_alertas or registros_cobrados:
        archivo_excel = exportar_excel(datos_alertas, registros_cobrados, prest_sin_pa, hoy)
        enviar_correo(archivo_excel)
    else:
        print("No hay registros relevantes para exportar.")

    cursor.close()
    conn.close()
    print("Conexión cerrada.")


if __name__ == "__main__":
    main()
